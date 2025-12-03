from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

IST = timezone(timedelta(hours=5, minutes=30))

logger = logging.getLogger(__name__)


HEADERS = [
    "VisitorID",
    "Name",
    "Phone",
    "Email",
    "Purpose",
    "ImagePath",
    "Timestamp",
    "Status",
]


@dataclass
class VisitorRecord:
    visitor_id: str
    name: str
    phone: str
    email: str
    purpose: str
    image_path: str
    timestamp: str
    status: str


class ExcelManager:
    def __init__(self, excel_path: Path) -> None:
        self.excel_path = excel_path
        self.excel_path.parent.mkdir(parents=True, exist_ok=True)
        if not self.excel_path.exists():
            self._create_workbook()

    def _create_workbook(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "VisitorLog"
        ws.append(HEADERS)
        wb.save(self.excel_path)

    def _load_workbook(self):
        if self.excel_path.exists():
            return load_workbook(self.excel_path)
        self._create_workbook()
        return load_workbook(self.excel_path)

    def _auto_adjust_columns(self, ws) -> None:
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    def generate_visitor_id(self) -> str:
        wb = self._load_workbook()
        ws = wb.active
        next_id = ws.max_row  # header is row 1
        wb.close()
        return f"VIS{next_id:05d}"

    def read_all_records(self) -> List[VisitorRecord]:
        wb = self._load_workbook()
        ws = wb.active
        records: List[VisitorRecord] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            records.append(VisitorRecord(*[str(x) if x is not None else "" for x in row]))
        wb.close()
        return records

    def check_suspicious(self, name: str, phone: str) -> Tuple[bool, List[int]]:
        today_str = datetime.now(IST).strftime("%Y-%m-%d")
        wb = self._load_workbook()
        ws = wb.active
        suspicious_rows: List[int] = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            r_name, r_phone, r_timestamp = row[1], row[2], row[6]
            if not r_name or not r_phone or not r_timestamp:
                continue
            try:
                ts_date = str(r_timestamp).split(" ")[0]
            except Exception:
                continue
            if r_name == name and r_phone == phone and ts_date == today_str:
                suspicious_rows.append(idx)
        wb.close()
        return (len(suspicious_rows) > 0, suspicious_rows)

    def append_record(self, record: VisitorRecord, suspicious_rows: List[int]) -> None:
        wb = self._load_workbook()
        ws = wb.active
        ws.append([
            record.visitor_id,
            record.name,
            record.phone,
            record.email,
            record.purpose,
            record.image_path,
            record.timestamp,
            record.status,
        ])

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row_idx in suspicious_rows:
            for cell in ws[row_idx]:
                cell.fill = yellow_fill

        self._auto_adjust_columns(ws)
        wb.save(self.excel_path)
        wb.close()
